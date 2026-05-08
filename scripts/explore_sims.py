import openpyxl

wb = openpyxl.load_workbook('./input/Simplified IMS Template.xlsm')
ws = wb['SIMS']

print(f'SIMS Sheet - Max Row: {ws.max_row}, Max Col: {ws.max_column}')
print()

# Print all rows with data
for row in range(1, ws.max_row + 1):
    cells = []
    for c in range(1, 35):
        v = ws.cell(row, c).value
        if v is not None:
            cells.append((c, v))
    if cells:
        print(f'Row {row}:')
        for c, v in cells:
            print(f'  C{c}={repr(v)}')
        print()
