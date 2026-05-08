import openpyxl

wb = openpyxl.load_workbook('./input/Simplified IMS Template.xlsm')
ws = wb['SIMS_Template']

print(f'Max Row: {ws.max_row}')
print(f'Max Col: {ws.max_column}')
print()

# Print merged cells
print('=== MERGED CELLS ===')
merged_list = list(ws.merged_cells.ranges)
for m in merged_list:
    print(f'  {m}')

print()

# Print all non-None cells row by row (first 40 rows, first 30 cols)
print('=== CELL VALUES (Rows 1-40, Cols 1-30) ===')
for row in range(1, 41):
    cells = []
    for c in range(1, 31):
        v = ws.cell(row, c).value
        if v is not None:
            cells.append(f'C{c}={repr(v)}')
    if cells:
        print(f'Row {row}: {" | ".join(cells)}')

print()

# Print all non-None cells row by row (first 40 rows, cols 30-60)
print('=== CELL VALUES (Rows 1-40, Cols 30-60) ===')
for row in range(1, 41):
    cells = []
    for c in range(30, 61):
        v = ws.cell(row, c).value
        if v is not None:
            cells.append(f'C{c}={repr(v)}')
    if cells:
        print(f'Row {row}: {" | ".join(cells)}')

print()

# Print all non-None cells row by row (first 40 rows, cols 60-100)
print('=== CELL VALUES (Rows 1-40, Cols 60-100) ===')
for row in range(1, 41):
    cells = []
    for c in range(60, 101):
        v = ws.cell(row, c).value
        if v is not None:
            cells.append(f'C{c}={repr(v)}')
    if cells:
        print(f'Row {row}: {" | ".join(cells)}')
