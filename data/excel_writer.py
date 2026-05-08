"""
ExcelWriter — Writes project updates back to the SIMS sheet.

SIMS Sheet Structure:
- Row 8: Headers
- Data rows alternate with "Durations in WD" rows (row 9, 11, 13, ...)
- Project data rows: 10, 12, 14, ...
- Each milestone has 3 columns:
  - {milestone} Baseline ECD
  - {milestone} Current ECD (CECD)
  - {milestone} Actual Completion Date (ACD)
"""

import openpyxl
import datetime
from typing import Dict
from data.models import Project, Milestone, Status
from config.settings import EXCEL_PATH


class ExcelWriter:
    """Write updated project data back to the Excel tracker (SIMS sheet)."""

    def __init__(self):
        pass

    def write_projects(self, projects: Dict[str, Project]):
        """
        Load the existing Excel workbook, update cells for each project
        in the SIMS sheet, and save back.
        """
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            sheet = wb['SIMS']
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found at: {EXCEL_PATH}")
        except KeyError:
            raise ValueError(f"Sheet 'SIMS' not found in {EXCEL_PATH}")

        # Row 8 is the header row
        header_row = 8
        headers = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(header_row, col).value
            if val is not None:
                headers[val] = col

        # Build reverse lookup: part_number -> row
        part_to_row = {}
        for row in range(header_row + 2, sheet.max_row + 1):
            part_val = sheet.cell(row, headers.get("Part Number", 3)).value
            ims_val = sheet.cell(row, headers.get("Simplified IMS", 6)).value
            
            # Skip "Durations in WD" rows
            if isinstance(ims_val, str) and "Durations" in str(ims_val):
                continue
                
            if part_val:
                part_to_row[str(part_val).strip()] = row

        for part_number, proj in projects.items():
            row = part_to_row.get(part_number)
            if row is None:
                continue

            # Update milestone columns
            for m_name, milestone in proj.milestones.items():
                # Actual Completion Date (ACD) column
                acd_header = f"{m_name} Actual Completion Date (ACD)"
                if acd_header in headers:
                    col = headers[acd_header]
                    if milestone.actual:
                        sheet.cell(row, col).value = milestone.actual

                # Current ECD (CECD) column - use for revised forecast
                cecd_header = f"{m_name} Current ECD (CECD)"
                if cecd_header in headers:
                    col = headers[cecd_header]
                    if milestone.revised_forecast:
                        if isinstance(milestone.revised_forecast, datetime.datetime):
                            sheet.cell(row, col).value = milestone.revised_forecast
                        else:
                            sheet.cell(row, col).value = milestone.revised_forecast

        wb.save(EXCEL_PATH)
