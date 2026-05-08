"""
ExcelLoader — Loads projects from the SIMS sheet.

SIMS Sheet Structure:
- Row 8: Headers (Project Name, AID, Part Number, Run Out Date, etc.)
- Row 9: "Durations in WD" (configuration row - skip)
- Row 10: Project data (e.g., "Advanced Manufacturing")
- Row 11: "Durations in WD" (skip)
- Row 12: Project data
- ...alternating pattern...

Each project row has:
- C1: Project Name
- C2: AID
- C3: Part Number
- C4: Run Out Date
- C5: Critical Part?
- C6: Simplified IMS type
- C7: Notes
- C8: Baseline Type
- C9+: Milestone columns (Baseline ECD, Current ECD, Actual Completion Date)

Each milestone has 3 columns:
  - {milestone} Baseline ECD
  - {milestone} Current ECD (CECD)
  - {milestone} Actual Completion Date (ACD)
"""

import openpyxl
from typing import Dict
from data.models import Project, Milestone, Person, Function, parse_date
from config.settings import EXCEL_PATH


class ExcelLoader:
    def __init__(self):
        pass

    def load_projects(self) -> Dict[str, Project]:
        """
        Loads projects from the SIMS sheet.
        Returns a dict keyed by part_number.
        """
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            sheet = wb['SIMS']
        except FileNotFoundError:
            raise FileNotFoundError(f"Excel file not found at: {EXCEL_PATH}")
        except KeyError:
            raise ValueError(f"Sheet 'SIMS' not found in {EXCEL_PATH}")

        # Row 8 is the header row
        header_row = 8
        headers = {}
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(header_row, col).value
            if val is not None:
                headers[val] = col

        col_name = headers.get("Project Name", 1)
        col_part = headers.get("Part Number", 3)
        col_ims = headers.get("Simplified IMS", 6)

        projects: Dict[str, Project] = {}
        total_rows_processed = 0

        # Data rows start from row 10 (alternating with "Durations in WD" rows)
        for row in range(header_row + 2, sheet.max_row + 1):
            name = sheet.cell(row, col_name).value
            part_number = sheet.cell(row, col_part).value
            ims_val = sheet.cell(row, col_ims).value

            if not name or not part_number:
                continue

            name_str = str(name).strip()

            # Skip "Durations in WD" rows (column F indicates duration config)
            if isinstance(ims_val, str) and "Durations" in str(ims_val):
                continue

            part_number = str(part_number).strip()

            # Extract project fields
            aid = sheet.cell(row, headers.get("AID", 2)).value
            run_out_date = sheet.cell(row, headers.get("Run Out Date", 4)).value
            critical = sheet.cell(row, headers.get("Critical Part?", 5)).value
            notes = sheet.cell(row, headers.get("Notes", 7)).value
            baseline_type = sheet.cell(row, headers.get("Baseline Type", 8)).value

            # Create Project
            proj = Project(
                name=name_str,
                part_number=part_number,
                program=str(aid) if aid else "",
                incoming_supplier="",
                outgoing_supplier="",
                line_down_date=str(run_out_date) if run_out_date else "",
                revenue_impact_daily=0.0,
            )

            # Load milestones from headers
            milestones_found = {}
            for header_name, col in headers.items():
                if " Baseline ECD" in header_name:
                    m_name = header_name.replace(" Baseline ECD", "").strip()
                    baseline_val = sheet.cell(row, col).value

                    # Find corresponding CECD and ACD columns
                    cecd_header = f"{m_name} Current ECD (CECD)"
                    acd_header = f"{m_name} Actual Completion Date (ACD)"

                    cecd_val = None
                    if cecd_header in headers:
                        cecd_val = sheet.cell(row, headers[cecd_header]).value

                    acd_val = None
                    if acd_header in headers:
                        acd_val = sheet.cell(row, headers[acd_header]).value

                    # Only add if there's actual data (not just formulas)
                    if baseline_val or acd_val:
                        milestones_found[m_name] = {
                            "baseline": str(baseline_val) if baseline_val else None,
                            "actual": str(acd_val) if acd_val else None,
                            "current_ecd": str(cecd_val) if cecd_val else None,
                        }

            # Add milestones to project
            for m_name, m_data in milestones_found.items():
                proj.add_milestone(
                    name=m_name,
                    baseline=m_data["baseline"],
                    actual=m_data["actual"],
                )
                # Store CECD as revised forecast
                if m_data["current_ecd"]:
                    cecd_date = parse_date(m_data["current_ecd"])
                    if cecd_date and m_name in proj.milestones:
                        proj.milestones[m_name].revised_forecast = cecd_date

            total_rows_processed += 1
            projects[part_number] = proj

        unique_count = len(projects)
        duplicates_removed = total_rows_processed - unique_count
        if duplicates_removed > 0:
            print(f"  [INFO] Deduplication: {total_rows_processed} rows loaded, {unique_count} unique projects ({duplicates_removed} duplicates removed)")

        return projects
